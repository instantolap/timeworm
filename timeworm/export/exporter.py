import csv
from datetime import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
)

from ..repository import TimeEntryRepository, quantize_hours


def _get_report_data(start_date: str, end_date: str,
                     project_id: Optional[int] = None,
                     customer_id: Optional[int] = None):
    entries = TimeEntryRepository.get_entries(
        project_id=project_id, customer_id=customer_id,
        start_date=start_date, end_date=end_date
    )
    summary = TimeEntryRepository.get_summary(
        start_date, end_date, customer_id=customer_id)
    customer_summary = TimeEntryRepository.get_summary_by_customer(
        start_date, end_date, customer_id=customer_id)
    return entries, summary, customer_summary


def export_csv(filepath: str, start_date: str, end_date: str,
               project_id: Optional[int] = None,
               customer_id: Optional[int] = None):
    """Export time entries as semicolon-delimited CSV."""
    entries, _, _ = _get_report_data(start_date, end_date, project_id, customer_id)
    with open(filepath, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f, delimiter=';')
        writer.writerow(['Kunde', 'Projekt', 'Datum', 'Start', 'Ende',
                         'Stunden', '€/h', 'Betrag', 'Beschreibung'])
        for e in entries:
            start = datetime.fromisoformat(e['start_time'])
            end = datetime.fromisoformat(e['end_time'])
            raw_hours = (end - start).total_seconds() / 3600
            q = e.get('time_quantum', 0.25) or 0.25
            hours = quantize_hours(raw_hours, q)
            amount = hours * e['hourly_rate']
            writer.writerow([
                e['customer_name'],
                e['project_name'],
                start.strftime('%d.%m.%Y'),
                start.strftime('%H:%M'),
                end.strftime('%H:%M'),
                f"{hours:.2f}",
                f"{e['hourly_rate']:.2f}",
                f"{amount:.2f}",
                e.get('description', ''),
            ])


def export_excel(filepath: str, start_date: str, end_date: str,
                 project_id: Optional[int] = None,
                 customer_id: Optional[int] = None):
    """Export time entries and summary as Excel workbook."""
    entries, summary, customer_summary = _get_report_data(start_date, end_date, project_id, customer_id)

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E3436", end_color="2E3436", fill_type="solid")

    # --- Detail sheet ---
    ws = wb.active
    ws.title = "Zeiteinträge"
    headers = ['Kunde', 'Projekt', 'Datum', 'Start', 'Ende',
               'Stunden', '€/h', 'Betrag (€)', 'Beschreibung']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row, e in enumerate(entries, 2):
        start = datetime.fromisoformat(e['start_time'])
        end = datetime.fromisoformat(e['end_time'])
        raw_hours = (end - start).total_seconds() / 3600
        q = e.get('time_quantum', 0.25) or 0.25
        hours = quantize_hours(raw_hours, q)
        amount = hours * e['hourly_rate']
        ws.cell(row=row, column=1, value=e['customer_name'])
        ws.cell(row=row, column=2, value=e['project_name'])
        ws.cell(row=row, column=3, value=start.strftime('%d.%m.%Y'))
        ws.cell(row=row, column=4, value=start.strftime('%H:%M'))
        ws.cell(row=row, column=5, value=end.strftime('%H:%M'))
        ws.cell(row=row, column=6, value=round(hours, 2))
        ws.cell(row=row, column=7, value=e['hourly_rate'])
        ws.cell(row=row, column=8, value=round(amount, 2))
        ws.cell(row=row, column=9, value=e.get('description', ''))

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 16

    # --- Summary sheet ---
    ws2 = wb.create_sheet("Zusammenfassung")
    sum_headers = ['Kunde', 'Projekt', 'Stunden', '€/h', 'Betrag (€)', 'Einträge']
    for col, h in enumerate(sum_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    total_hours = 0
    total_amount = 0
    row_num = 2

    # Group by customer
    projects_by_customer = {}
    for s in summary:
        projects_by_customer.setdefault(s['customer_name'], []).append(s)

    for cust in customer_summary:
        cname = cust['customer_name']
        chours = cust['total_hours'] or 0
        camount = cust['total_amount'] or 0
        total_hours += chours
        total_amount += camount

        # Customer row (bold)
        ws2.cell(row=row_num, column=1, value=cname).font = Font(bold=True)
        ws2.cell(row=row_num, column=3, value=round(chours, 2)).font = Font(bold=True)
        ws2.cell(row=row_num, column=5, value=round(camount, 2)).font = Font(bold=True)
        row_num += 1

        for proj in projects_by_customer.get(cname, []):
            phours = proj['total_hours'] or 0
            pamount = phours * proj['hourly_rate']
            ws2.cell(row=row_num, column=2, value=proj['project_name'])
            ws2.cell(row=row_num, column=3, value=round(phours, 2))
            ws2.cell(row=row_num, column=4, value=proj['hourly_rate'])
            ws2.cell(row=row_num, column=5, value=round(pamount, 2))
            ws2.cell(row=row_num, column=6, value=proj['entry_count'])
            row_num += 1

    # Totals
    ws2.cell(row=row_num, column=1, value="GESAMT").font = Font(bold=True, size=12)
    ws2.cell(row=row_num, column=3, value=round(total_hours, 2)).font = Font(bold=True, size=12)
    ws2.cell(row=row_num, column=5, value=round(total_amount, 2)).font = Font(bold=True, size=12)

    for col in ws2.columns:
        ws2.column_dimensions[col[0].column_letter].width = 18

    wb.save(filepath)


def export_pdf(filepath: str, start_date: str, end_date: str,
               project_id: Optional[int] = None,
               customer_id: Optional[int] = None):
    """Export as PDF Stundennachweis."""
    entries, summary, customer_summary = _get_report_data(start_date, end_date, project_id, customer_id)

    doc = SimpleDocTemplate(
        filepath, pagesize=A4,
        leftMargin=20 * mm, rightMargin=20 * mm,
        topMargin=20 * mm, bottomMargin=20 * mm
    )
    styles = getSampleStyleSheet()
    elements = []

    # Title
    title_style = ParagraphStyle(
        'InvoiceTitle', parent=styles['Title'],
        fontSize=18, spaceAfter=6 * mm
    )
    elements.append(Paragraph("Stundennachweis", title_style))

    # Period — end_date is exclusive (first of next month), show last day of range
    start_dt = datetime.fromisoformat(start_date)
    end_dt = datetime.fromisoformat(end_date)
    from datetime import timedelta
    last_day = end_dt - timedelta(days=1)
    period = (
        f"Zeitraum: {start_dt.strftime('%d.%m.%Y')} \u2013 "
        f"{last_day.strftime('%d.%m.%Y')}"
    )
    elements.append(Paragraph(period, styles['Normal']))
    elements.append(Spacer(1, 8 * mm))

    # --- Summary by customer ---
    if customer_summary:
        elements.append(Paragraph("Zusammenfassung nach Kunde", styles['Heading2']))

        sum_cell = ParagraphStyle('SumCell', parent=styles['Normal'],
                                   fontSize=8, leading=10)
        sum_cell_bold = ParagraphStyle('SumCellBold', parent=sum_cell,
                                        fontName='Helvetica-Bold')
        sum_hdr = ParagraphStyle('SumHdr', parent=sum_cell_bold,
                                  textColor=colors.white)

        sum_data = [[Paragraph('Kunde / Projekt', sum_hdr),
                     Paragraph('Stunden', sum_hdr),
                     Paragraph('\u20ac/h', sum_hdr),
                     Paragraph('Betrag (\u20ac)', sum_hdr)]]
        total_hours = 0.0
        total_amount = 0.0

        projects_by_customer = {}
        for s in summary:
            projects_by_customer.setdefault(s['customer_name'], []).append(s)

        for cust in customer_summary:
            cname = cust['customer_name']
            chours = cust['total_hours'] or 0
            camount = cust['total_amount'] or 0
            total_hours += chours
            total_amount += camount
            sum_data.append([Paragraph(cname, sum_cell_bold), Paragraph(f"{chours:.2f}", sum_cell_bold), Paragraph('', sum_cell), Paragraph(f"{camount:.2f}", sum_cell_bold)])

            for proj in projects_by_customer.get(cname, []):
                phours = proj['total_hours'] or 0
                pamount = phours * proj['hourly_rate']
                sum_data.append([
                    Paragraph(f"    {proj['project_name']}", sum_cell),
                    Paragraph(f"{phours:.2f}", sum_cell),
                    Paragraph(f"{proj['hourly_rate']:.2f}", sum_cell),
                    Paragraph(f"{pamount:.2f}", sum_cell),
                ])

        sum_data.append([Paragraph('Gesamt', sum_cell_bold), Paragraph(f"{total_hours:.2f}", sum_cell_bold), Paragraph('', sum_cell), Paragraph(f"{total_amount:.2f}", sum_cell_bold)])

        # Full page width: A4=210mm - 2*20mm margins = 170mm
        page_w = 170 * mm
        col_widths = [page_w * 0.35, page_w * 0.18, page_w * 0.15, page_w * 0.32]
        sum_table = Table(sum_data, colWidths=col_widths)
        sum_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E3436')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2),
             [colors.white, colors.HexColor('#F5F5F5')]),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        elements.append(sum_table)
        elements.append(Spacer(1, 8 * mm))

    # --- Detail table ---
    if entries:
        elements.append(Paragraph("Einzelnachweise", styles['Heading2']))

        # Cell style for wrapping text
        cell_style = ParagraphStyle('CellWrap', parent=styles['Normal'],
                                     fontSize=7, leading=9)
        cell_style_bold = ParagraphStyle('CellWrapBold', parent=cell_style,
                                          fontName='Helvetica-Bold',
                                          textColor=colors.white)

        detail_data = [[
            Paragraph('Kunde', cell_style_bold),
            Paragraph('Projekt', cell_style_bold),
            Paragraph('Datum', cell_style_bold),
            Paragraph('Von', cell_style_bold),
            Paragraph('Bis', cell_style_bold),
            Paragraph('Stunden', cell_style_bold),
            Paragraph('Betrag (€)', cell_style_bold),
            Paragraph('Beschreibung', cell_style_bold),
        ]]
        for e in entries:
            start = datetime.fromisoformat(e['start_time'])
            end = datetime.fromisoformat(e['end_time'])
            raw_hours = (end - start).total_seconds() / 3600
            q = e.get('time_quantum', 0.25) or 0.25
            hours = quantize_hours(raw_hours, q)
            amount = hours * e['hourly_rate']
            detail_data.append([
                Paragraph(e['customer_name'], cell_style),
                Paragraph(e['project_name'], cell_style),
                Paragraph(start.strftime('%d.%m.%Y'), cell_style),
                Paragraph(start.strftime('%H:%M'), cell_style),
                Paragraph(end.strftime('%H:%M'), cell_style),
                Paragraph(f"{hours:.2f}", cell_style),
                Paragraph(f"{amount:.2f}", cell_style),
                Paragraph(e.get('description', '') or '', cell_style),
            ])

        # Full page width: A4=210mm - 2*20mm margins = 170mm
        # Fixed cols: Datum(18mm) Von(12mm) Bis(12mm) Stunden(14mm) Betrag(18mm) = 74mm
        # Flexible: Kunde + Projekt + Beschreibung share remaining 96mm
        page_w = 170 * mm
        fixed = 74 * mm
        flex = page_w - fixed
        col_widths = [flex * 0.28, flex * 0.28, 18*mm, 12*mm, 12*mm,
                      14*mm, 18*mm, flex * 0.44]
        detail_table = Table(detail_data, colWidths=col_widths, repeatRows=1)
        detail_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E3436')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (3, 0), (6, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1),
             [colors.white, colors.HexColor('#F5F5F5')]),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        elements.append(detail_table)

    doc.build(elements)
